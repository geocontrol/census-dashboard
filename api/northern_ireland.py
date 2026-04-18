"""
Northern Ireland Census 2021 — Data Zone (DZ2021) boundaries and
population data (MS-A01).

NISRA publishes DZ2021 boundaries directly as GeoJSON in WGS84 (CRS84),
so no reprojection is required (unlike Scotland's BNG shapefiles). The
MS-A01 usual-resident-population table is published as XLSX with a
clean DZ-level sheet, so no OA→DZ aggregation is needed either.

Indicator ingestion from NISRA cross-tab CSVs is a separate sub-task
and is intentionally not implemented here yet.
"""
import json
import logging
import zipfile
from pathlib import Path
from typing import Optional

import httpx
from openpyxl import load_workbook
from shapely.geometry import MultiPolygon, Polygon, mapping, shape

logger = logging.getLogger(__name__)

DZ2021_BOUNDARY_GEOJSON_URL = (
    "https://www.nisra.gov.uk/files/nisra/publications/geography-dz2021-geojson.zip"
)
DZ2021_LOOKUP_XLSX_URL = (
    "https://www.nisra.gov.uk/files/nisra/documents/2025-04/"
    "geography-data-zone-and-super-data-zone-lookups-v3.xlsx"
)
MS_A01_URL = "https://www.nisra.gov.uk/system/files/statistics/census-2021-ms-a01.xlsx"

# Bulk Main Statistics zips — each contains all phase tables at LGD/Ward/Settlement/NI level
MS_BULK_URLS = {
    "phase1": "https://www.nisra.gov.uk/system/files/statistics/census-2021-main-statistics-for-northern-ireland-phase-1-all-tables.zip",
    "phase2": "https://www.nisra.gov.uk/system/files/statistics/census-2021-main-statistics-for-northern-ireland-phase-2-all-tables.zip",
    "phase3": "https://www.nisra.gov.uk/system/files/statistics/census-2021-main-statistics-for-northern-ireland-phase-3-all-tables.zip",
}

TARGET_VERTICES = 30


def _simplify_ring(coords: list, target: int = TARGET_VERTICES) -> list:
    n = len(coords)
    if n <= target:
        return coords
    step = max(1, n // target)
    simplified = coords[::step]
    if simplified[0] != simplified[-1]:
        simplified.append(simplified[0])
    return simplified


def _simplify_geometry(geom):
    if geom.geom_type == "Polygon":
        exterior = _simplify_ring(list(geom.exterior.coords))
        interiors = [_simplify_ring(list(ring.coords)) for ring in geom.interiors]
        try:
            return Polygon(exterior, interiors)
        except Exception:
            return geom
    if geom.geom_type == "MultiPolygon":
        parts = []
        for part in geom.geoms:
            simplified = _simplify_geometry(part)
            if simplified.is_valid and not simplified.is_empty:
                parts.append(simplified)
        if parts:
            return MultiPolygon(parts)
    return geom


async def download_dz21_boundaries(data_dir: Path) -> Path:
    """Download NISRA DZ2021 GeoJSON, simplify, normalise properties, save."""
    output = data_dir / "boundaries_ni_dz21.geojson"
    if output.exists():
        logger.info(
            f"NI DZ2021 boundaries cached ({output.stat().st_size / 1024 / 1024:.1f} MB)"
        )
        return output

    logger.info("Downloading NI DZ2021 boundaries (~26 MB)...")
    zip_path = data_dir / "geography-dz2021-geojson.zip"

    async with httpx.AsyncClient(timeout=120.0, follow_redirects=True) as client:
        resp = await client.get(DZ2021_BOUNDARY_GEOJSON_URL)
        resp.raise_for_status()
        zip_path.write_bytes(resp.content)
        logger.info(f"  Downloaded {zip_path.stat().st_size / 1024 / 1024:.1f} MB")

    with zipfile.ZipFile(zip_path) as zf:
        geojson_name = next(
            (n for n in zf.namelist() if n.lower().endswith(".geojson")), None
        )
        if not geojson_name:
            raise FileNotFoundError("No .geojson file found in DZ2021 boundary zip")
        with zf.open(geojson_name) as f:
            raw = json.loads(f.read().decode("utf-8"))

    features = []
    for feat in raw.get("features", []):
        props = feat.get("properties", {}) or {}
        dz_code = props.get("DZ2021_cd") or props.get("DZ2021_CD")
        dz_name = props.get("DZ2021_nm") or props.get("DZ2021_NM") or dz_code
        if not dz_code:
            continue

        try:
            geom = shape(feat["geometry"])
        except Exception:
            continue
        if geom.is_empty:
            continue
        if not geom.is_valid:
            geom = geom.buffer(0)

        simplified = _simplify_geometry(geom)
        if not simplified.is_valid:
            simplified = simplified.buffer(0)

        try:
            area_ha = float(props.get("Area_ha") or 0)
        except (TypeError, ValueError):
            area_ha = 0.0
        area_km2 = round(area_ha / 100.0, 4) if area_ha else 0.0

        features.append(
            {
                "type": "Feature",
                "properties": {
                    "DZ2021CD": dz_code,
                    "DZ2021NM": dz_name,
                    "SDZ2021CD": props.get("SDZ2021_cd", ""),
                    "SDZ2021NM": props.get("SDZ2021_nm", ""),
                    "LGD2014CD": props.get("LGD2014_cd", ""),
                    "LGD2014NM": props.get("LGD2014_nm", ""),
                    "nation": "NI",
                    "area_km2": area_km2,
                },
                "geometry": mapping(simplified),
            }
        )

    geojson = {"type": "FeatureCollection", "features": features}
    output.write_text(json.dumps(geojson))
    logger.info(
        f"NI DZ2021 boundaries: {len(features)} features, "
        f"{output.stat().st_size / 1024 / 1024:.1f} MB"
    )

    try:
        zip_path.unlink()
    except Exception:
        pass

    return output


# ═══════ Northern Ireland Local Government Districts (LGD2014) ═══════
# Sourced from NISRA DZ2021 lookup table (geography-data-zone-and-super-data-zone-lookups-v3.xlsx).
# These 11 LGDs are the equivalent of Scotland's Council Areas / E&W's LADs.

NI_LOCAL_GOVERNMENT_DISTRICTS = [
    {"code": "N09000001", "name": "Antrim and Newtownabbey"},
    {"code": "N09000002", "name": "Armagh City, Banbridge and Craigavon"},
    {"code": "N09000003", "name": "Belfast"},
    {"code": "N09000004", "name": "Causeway Coast and Glens"},
    {"code": "N09000005", "name": "Derry City and Strabane"},
    {"code": "N09000006", "name": "Fermanagh and Omagh"},
    {"code": "N09000007", "name": "Lisburn and Castlereagh"},
    {"code": "N09000008", "name": "Mid and East Antrim"},
    {"code": "N09000009", "name": "Mid Ulster"},
    {"code": "N09000010", "name": "Newry, Mourne and Down"},
    {"code": "N09000011", "name": "Ards and North Down"},
]


def get_ni_lgd_dzs(lgd_code: str, dz21_geojson_path: Path) -> list[str]:
    """Return all DZ2021 codes inside a Northern Ireland Local Government District."""
    geojson = json.loads(dz21_geojson_path.read_text())
    return [
        feat["properties"]["DZ2021CD"]
        for feat in geojson.get("features", [])
        if feat.get("properties", {}).get("LGD2014CD") == lgd_code
        and feat.get("properties", {}).get("DZ2021CD")
    ]


# ═══════ Indicator mapping ═══════
# Maps E&W dataset IDs to a NISRA Main Statistics table + column-extraction
# rules for the LGD sheet. Each NI Data Zone inherits the value of its
# parent Local Government District.
#
# Schema:
#   "table":            MS table code, e.g. "D01" → census-2021-ms-d01.xlsx
#   "scope_prefix":     optional — only consider columns whose header (case-
#                       insensitive, normalised whitespace) starts with this
#                       prefix. Used to scope to top-level totals in cross-tabs.
#   "numerator":        list of substrings; columns whose header contains any
#                       of them (within scope) are summed for the numerator.
#                       Columns are deduplicated by index, so overlapping
#                       substrings (e.g. "Good health" matching both "Good
#                       health" and "Very good health") are counted once.
#   "denominator":      optional — list of substrings used as the denominator
#                       sum. If omitted, the per-LGD total in column 2 is used.
#   "exact":            optional bool, default False. When True, a header
#                       matches a needle only if the normalised header EQUALS
#                       the normalised needle. Use for column labels that
#                       overlap (e.g. "Married" vs "Single (never married…)").

NI_INDICATOR_MAP: dict = {
    # ── Health (MS-D01) ──
    "health_good": {
        "table": "D01",
        "scope_prefix": "All usual residents:",
        "numerator": ["Very good health", "Good health"],
    },
    "health_bad": {
        "table": "D01",
        "scope_prefix": "All usual residents:",
        "numerator": ["Bad health", "Very bad health"],
    },
    # ── Disability (MS-D02) ──
    "disability_limited_lot": {
        "table": "D02",
        "scope_prefix": "All usual residents:",
        "numerator": ["Day-to-day activities limited a lot"],
    },
    # ── Unpaid care (MS-D17) ──
    "unpaid_care": {
        "table": "D17",
        "scope_prefix": "All usual residents aged 5 and over:",
        "numerator": [
            "Provides 1-19 hours unpaid care",
            "Provides 20-34 hours unpaid care",
            "Provides 35-49 hours unpaid care",
            "Provides 50+ hours unpaid care",
        ],
    },
    # ── Housing tenure (MS-E15 — households) ──
    "home_ownership": {"table": "E15", "numerator": ["Owner occupied:"]},
    "social_rented":  {"table": "E15", "numerator": ["Social rented:"]},
    "private_rented": {"table": "E15", "numerator": ["Private rented:"]},
    # ── Accommodation (MS-E06) ──
    "accommodation_detached": {"table": "E06", "numerator": ["Detached"]},
    "accommodation_terraced": {"table": "E06", "numerator": ["Terraced"]},
    "accommodation_flat":     {"table": "E06", "numerator": ["Flat, maisonette or apartment"]},
    # ── Heating (MS-E11 — household-based) ──
    "no_central_heating": {"table": "E11", "numerator": ["No central heating"]},
    "gas_heating":        {"table": "E11", "numerator": ["Mains gas only"]},
    "electric_heating":   {"table": "E11", "numerator": ["Electric (for example storage heaters) only"]},
    # ── Cars (MS-E10) ──
    "car_none": {"table": "E10", "numerator": ["No cars or vans available"]},
    # ── Travel to work (MS-I01) ──
    "travel_car":       {"table": "I01", "numerator": ["Driving a car or van"]},
    "travel_public":    {"table": "I01", "numerator": ["Bus, minibus or coach", "Train"]},
    "work_from_home":   {"table": "I01", "numerator": ["Work mainly at or from home"]},
    "travel_walk_cycle": {"table": "I01", "numerator": ["Bicycle", "On foot"]},
    # ── Qualifications (MS-G01 — aged 16+) ──
    "qualifications_level4": {"table": "G01", "numerator": ["Level 4 qualifications and above"]},
    "no_qualifications":     {"table": "G01", "numerator": ["No qualifications"]},
    # ── Economy (MS-H02) ──
    "economic_activity": {
        "table": "H02",
        "scope_prefix": "Usual residents aged 16 and over:",
        "numerator": ["Economically active:"],
    },
    "unemployment": {
        "table": "H02",
        "scope_prefix": "Usual residents aged 16 and over:",
        "numerator": ["Economically active: Unemployed"],
        "denominator": ["Economically active:"],
    },
    # ── Sex (MS-A07) ──
    "sex_female": {"table": "A07", "numerator": ["Female"]},
    # ── Country of birth (MS-A16) ──
    "born_uk": {
        "table": "A16",
        "numerator": [
            "United Kingdom:\n Northern Ireland",
            "United Kingdom:\n England",
            "United Kingdom:\n Scotland",
            "United Kingdom:\n Wales",
        ],
    },
    # ── National identity (MS-B15) ──
    "identity_british_only": {"table": "B15", "numerator": ["British only"]},
    # ── Ethnicity (MS-B01) ──
    "white_british": {"table": "B01", "numerator": ["White"]},
    "ethnic_asian":  {"table": "B01", "numerator": [
        "Indian", "Chinese", "Filipino", "Pakistani", "Other Asian",
    ]},
    "ethnic_black":  {"table": "B01", "numerator": ["Black African", "Black Other"]},
    "ethnic_mixed":  {"table": "B01", "numerator": ["Mixed"]},
    # ── Religion (MS-B19) ──
    "christian":   {"table": "B19", "numerator": [
        "Catholic", "Presbyterian Church in Ireland", "Church of Ireland",
        "Methodist Church in Ireland", "Other Christian",
    ]},
    "no_religion": {"table": "B19", "numerator": ["No religion"]},
    # ── Marital status (MS-A30) — exact match because "Married" is also a
    # substring of "Single (never married…)" and "Separated (but still legally
    # married…)".
    "married":       {"table": "A30", "exact": True, "numerator": ["Married", "In a civil partnership"]},
    "never_married": {"table": "A30", "numerator": ["Single (never married"]},
    "divorced":      {"table": "A30", "numerator": ["Divorced or formerly in a civil partnership"]},
    "widowed":       {"table": "A30", "numerator": ["Widowed or surviving partner"]},
    # ── Households (MS-A26) ──
    "hh_one_person":     {"table": "A26", "numerator": ["One person household"]},
    "hh_lone_parent":    {"table": "A26", "numerator": ["Lone parent family"]},
    "hh_married_couple": {"table": "A26", "numerator": ["Married or civil partnership couple"]},
}


# ═══════ Population data (MS-A01) ═══════


async def download_ni_population_xlsx(data_dir: Path) -> Path:
    """Download NISRA MS-A01 (Usual resident population by Data Zone)."""
    output = data_dir / "ni_ms_a01.xlsx"
    if output.exists():
        logger.info(f"NI MS-A01 cached ({output.stat().st_size / 1024:.0f} KB)")
        return output

    logger.info("Downloading NI MS-A01 population (~470 KB)...")
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        resp = await client.get(MS_A01_URL)
        resp.raise_for_status()
        output.write_bytes(resp.content)
        logger.info(f"  Downloaded {output.stat().st_size / 1024:.0f} KB")
    return output


def parse_ni_population_dz(xlsx_path: Path) -> dict[str, int]:
    """Parse the DZ sheet of MS-A01. Returns {dz_code: usual_residents}."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["DZ"]

    populations: dict[str, int] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or len(row) < 3:
            continue
        code = (row[1] or "").strip() if isinstance(row[1], str) else None
        if not code or not code.startswith("N20"):
            continue
        try:
            populations[code] = int(row[2])
        except (TypeError, ValueError):
            continue
    return populations


def _stats(values: dict[str, float]) -> dict:
    sv = sorted(values.values())
    n = len(sv)
    if n == 0:
        return {}
    return {
        "min": sv[0], "max": sv[-1],
        "mean": round(sum(sv) / n, 2),
        "p10": sv[int(n * 0.1)], "p25": sv[int(n * 0.25)],
        "p50": sv[int(n * 0.5)], "p75": sv[int(n * 0.75)],
        "p90": sv[int(n * 0.9)], "count": n,
    }


def compute_ni_population_data(
    dz21_geojson_path: Path,
    ms_a01_xlsx_path: Path,
) -> dict[str, dict]:
    """Build population_total + population_density payloads for NI Data Zones.

    Population comes from MS-A01; area_km2 comes from the boundary GeoJSON
    (already populated during boundary ingestion).
    """
    geojson = json.loads(dz21_geojson_path.read_text())
    areas: dict[str, float] = {}
    names: dict[str, str] = {}
    for feat in geojson.get("features", []):
        props = feat.get("properties", {}) or {}
        code = props.get("DZ2021CD")
        if not code:
            continue
        names[code] = props.get("DZ2021NM", code)
        try:
            areas[code] = float(props.get("area_km2") or 0.0)
        except (TypeError, ValueError):
            areas[code] = 0.0

    populations = parse_ni_population_dz(ms_a01_xlsx_path)

    total_values: dict[str, float] = {}
    density_values: dict[str, float] = {}
    for code, pop in populations.items():
        if code not in areas:
            continue
        total_values[code] = float(pop)
        area = areas[code]
        density_values[code] = round(pop / area, 2) if area > 0 else 0.0

    source = "NISRA Census 2021 — MS-A01 (population) + DZ2021 boundary area"
    return {
        "population_total": {
            "dataset_id": "population_total",
            "values": total_values,
            "names": {c: names.get(c, c) for c in total_values},
            "stats": _stats(total_values),
            "source": source,
        },
        "population_density": {
            "dataset_id": "population_density",
            "values": density_values,
            "names": {c: names.get(c, c) for c in density_values},
            "stats": _stats(density_values),
            "source": source,
        },
    }


def get_ni_data_for_dataset(
    dataset_id: str,
    data_dir: Path,
) -> Optional[dict]:
    """Load cached NI data payload for a dataset, if present."""
    cache_file = data_dir / f"data_{dataset_id}_national_ni.json"
    if cache_file.exists():
        return json.loads(cache_file.read_text())
    return None


# ═══════ Main Statistics — LGD-level indicator ingestion ═══════


async def download_ni_main_statistics(data_dir: Path) -> Path:
    """Download and extract NISRA Main Statistics phase 1/2/3 zips.

    Returns the directory containing all extracted MS-*.xlsx files.
    Cached on disk; subsequent calls are no-ops.
    """
    out_dir = data_dir / "ni_main_statistics"
    marker = out_dir / ".downloaded"
    if marker.exists():
        logger.info("NI Main Statistics tables cached")
        return out_dir

    out_dir.mkdir(exist_ok=True)
    async with httpx.AsyncClient(timeout=300.0, follow_redirects=True) as client:
        for phase, url in MS_BULK_URLS.items():
            logger.info(f"Downloading NI Main Statistics {phase}...")
            resp = await client.get(url)
            resp.raise_for_status()
            zip_path = out_dir / f"{phase}.zip"
            zip_path.write_bytes(resp.content)
            logger.info(f"  {phase}: {zip_path.stat().st_size / 1024 / 1024:.1f} MB")
            with zipfile.ZipFile(zip_path) as zf:
                zf.extractall(out_dir)
            zip_path.unlink()

    marker.touch()
    return out_dir


def _normalise(s: str) -> str:
    return " ".join((s or "").split()).lower()


def parse_ni_lgd_table(xlsx_path: Path) -> tuple[list[str], dict[str, list[float]]]:
    """Parse the LGD sheet of a NISRA MS table.

    Returns (header_row, rows_by_lgd_code) where rows_by_lgd_code is
    {N09…: [cell values, parallel to header_row]}.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = "LGD" if "LGD" in wb.sheetnames else None
    if not sheet_name:
        return [], {}

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    data_start = None
    for i, r in enumerate(rows):
        if len(r) > 1 and isinstance(r[1], str) and r[1].startswith("N09"):
            data_start = i
            break
    if data_start is None or data_start == 0:
        return [], {}

    header = [str(h) if h is not None else "" for h in rows[data_start - 1]]

    data: dict[str, list[float]] = {}
    for r in rows[data_start:]:
        if not (len(r) > 1 and isinstance(r[1], str) and r[1].startswith("N09")):
            break  # blank row separates count and percentage tables
        code = r[1].strip()
        values: list[float] = []
        for cell in r:
            if isinstance(cell, (int, float)):
                values.append(float(cell))
            else:
                values.append(0.0)
        data[code] = values
    return header, data


def _resolve_columns(
    header: list[str],
    needles: list[str],
    scope_prefix: Optional[str],
    exact: bool = False,
) -> list[int]:
    """Return indices of columns whose header matches any needle within scope.

    Matching is case-insensitive on whitespace-normalised text. With
    exact=False, a header matches if it CONTAINS the needle as a substring.
    With exact=True, the header (after stripping any scope_prefix) must EQUAL
    the needle. Indices are deduplicated, so overlapping needles count their
    column once.
    """
    scope_norm = _normalise(scope_prefix) if scope_prefix else None
    needles_norm = [_normalise(n) for n in needles]

    matched: set[int] = set()
    for i, h in enumerate(header):
        if i < 2:  # skip Geography + Geography code columns
            continue
        h_norm = _normalise(h)
        if scope_norm:
            if not h_norm.startswith(scope_norm):
                continue
            payload = h_norm[len(scope_norm):].lstrip()
        else:
            payload = h_norm
        for n in needles_norm:
            if not n:
                continue
            if exact:
                if payload == n:
                    matched.add(i)
                    break
            else:
                if n in h_norm:
                    matched.add(i)
                    break
    return sorted(matched)


def _ms_table_path(ms_dir: Path, table_code: str) -> Optional[Path]:
    fname = f"census-2021-ms-{table_code.lower()}.xlsx"
    candidate = ms_dir / fname
    return candidate if candidate.exists() else None


def process_ni_indicator(
    dataset_id: str,
    ms_dir: Path,
    dz21_geojson_path: Path,
) -> Optional[dict]:
    """Process a single NI indicator from MS bulk tables.

    Returns the standard {dataset_id, values, names, stats, source} payload
    where each NI Data Zone inherits its parent LGD's rate.
    """
    config = NI_INDICATOR_MAP.get(dataset_id)
    if not config:
        return None

    table_code = config["table"]
    xlsx_path = _ms_table_path(ms_dir, table_code)
    if not xlsx_path:
        logger.warning(f"NI MS table not found for {dataset_id}: MS-{table_code}")
        return None

    header, rows = parse_ni_lgd_table(xlsx_path)
    if not rows:
        logger.warning(f"NI {dataset_id}: no LGD rows parsed from MS-{table_code}")
        return None

    exact = bool(config.get("exact"))
    num_cols = _resolve_columns(header, config["numerator"], config.get("scope_prefix"), exact=exact)
    if not num_cols:
        logger.warning(
            f"NI {dataset_id}: no numerator columns matched in MS-{table_code} "
            f"(needles={config['numerator']}, scope={config.get('scope_prefix')!r}, exact={exact})"
        )
        return None

    den_cols = _resolve_columns(
        header, config["denominator"], config.get("scope_prefix"), exact=exact
    ) if config.get("denominator") else None

    lgd_rates: dict[str, float] = {}
    for code, values in rows.items():
        num = sum(values[c] for c in num_cols if c < len(values))
        if den_cols is not None:
            den = sum(values[c] for c in den_cols if c < len(values))
        else:
            den = values[2] if len(values) > 2 else 0
        if den > 0:
            lgd_rates[code] = round(num / den * 100, 2)
        else:
            lgd_rates[code] = 0.0

    # Inherit each DZ's rate from its parent LGD
    geojson = json.loads(dz21_geojson_path.read_text())
    dz_values: dict[str, float] = {}
    dz_names: dict[str, str] = {}
    for feat in geojson.get("features", []):
        props = feat.get("properties", {}) or {}
        dz = props.get("DZ2021CD")
        lgd = props.get("LGD2014CD")
        if not dz or not lgd:
            continue
        if lgd in lgd_rates:
            dz_values[dz] = lgd_rates[lgd]
            dz_names[dz] = props.get("DZ2021NM", dz)

    if not dz_values:
        return None

    return {
        "dataset_id": dataset_id,
        "values": dz_values,
        "names": dz_names,
        "stats": _stats(dz_values),
        "source": (
            f"NISRA Census 2021 — MS-{table_code} (LGD-level; each Data Zone "
            f"shows its parent Local Government District's rate)"
        ),
        "granularity": "lgd",
    }


def process_all_ni_indicators(
    ms_dir: Path,
    dz21_geojson_path: Path,
) -> dict[str, dict]:
    """Process every indicator in NI_INDICATOR_MAP. Returns {dataset_id: payload}."""
    results: dict[str, dict] = {}
    for dataset_id in NI_INDICATOR_MAP:
        result = process_ni_indicator(dataset_id, ms_dir, dz21_geojson_path)
        if result:
            results[dataset_id] = result
    return results
